# # # import requests 
# # # from bs4 import BeautifulSoup 
# # # import csv 
# # # from pprint import pprint
# # # import bs4 as bs


   
# # # URL = "https://in.tradingview.com/symbols/NSE-POWERGRID/"
# # # r = requests.get(URL, headers={"User-Agent":"Mozilla/5.0"}) 

# # # soup = BeautifulSoup(r.content, 'html5lib') 
    
# # # table = soup.find_all('div', attrs = {'class':'tv-widget-fundamentals__row'}) 

# # # for i in table:
# # #     print(i)
# # # Library for opening url and creating 
# # # requests 
# # import urllib.request 

# # # pretty-print python data structures 
# # from pprint import pprint 

# # # for parsing all the tables present 
# # # on the website 
# # from html_table_parser import HTMLTableParser 

# # # for converting the parsed data in a 
# # # pandas dataframe 
# # import pandas as pd 


# # # Opens a website and read its 
# # # binary contents (HTTP Response Body) 
# # def url_get_contents(url): 

# # 	# Opens a website and read its 
# # 	# binary contents (HTTP Response Body) 

# # 	#making request to the website 
# # 	req = urllib.request.Request(url=url) 
# # 	f = urllib.request.urlopen(req) 

# # 	#reading contents of the website 
# # 	return f.read() 

# # # defining the html contents of a URL. 
# # xhtml = url_get_contents('https://www.moneycontrol.com/india/stockpricequote/power-generationdistribution/powergridcorporationindia/PGC').decode('utf-8') 

# # # Defining the HTMLTableParser object 
# # p = HTMLTableParser() 

# # # feeding the html contents in the 
# # # HTMLTableParser object 
# # p.feed(xhtml) 

# # # Now finally obtaining the data of 
# # # the table required 
# # pprint(p.tables) 

# # # converting the parsed data to 
# # # datframe 
# # print("\n\nPANDAS DATAFRAME\n") 
# # print(pd.DataFrame(p.tables[1]))
# # def render(source_html):
# #     """Fully render HTML, JavaScript and all."""

# #     import sys
# #     from PyQt5.QtWidgets import QApplication
# #     from PyQt5.QtWebEngineWidgets import QWebEngineView

# #     class Render(QWebEngineView):
# #         def __init__(self, html):
# #             self.html = None
# #             self.app = QApplication(sys.argv)
# #             QWebEngineView.__init__(self)
# #             self.loadFinished.connect(self._loadFinished)
# #             self.setHtml(html)
# #             self.app.exec_()

# #         def _loadFinished(self, result):
# #             # This is an async call, you need to wait for this
# #             # to be called before closing the app
# #             self.page().toHtml(self.callable)

# #         def callable(self, data):
# #             self.html = data
# #             # Data has been stored, it's safe to quit the app
# #             self.app.quit()

# #     return Render(source_html).html


# # import requests
# # sample_html = requests.get('https://in.tradingview.com/symbols/NSE-MRPL/').text
# # soup = render(sample_html)

# # print(soup)
# # if 'Book' in soup:
# #     print('sd')



# '''
# # import requests 
# # # from bs4 import BeautifulSoup 
# # import csv 
# # from pprint import pprint
# import bs4 as bs

# import sys
# import time
# from PyQt5.QtWebEngineWidgets import QWebEnginePage
# from PyQt5.QtWidgets import QApplication
# from PyQt5.QtCore import QUrl
# print('sdsd')

# class Page(QWebEnginePage):
#     def __init__(self, url):
#         self.app = QApplication(sys.argv)
#         QWebEnginePage.__init__(self)
#         self.html = ''
#         self.loadFinished.connect(self.on_load_finished)
#         self.load(QUrl(url))
#         self.triggerAction(QWebEnginePage.ReloadAndBypassCache, True)
#         self.app.exec_()

#     def on_load_finished(self):
#         self.html = self.toHtml(self.Callable)
#         print('Load finished')

#     def Callable(self, html_str):
#         self.html = html_str
#         self.app.quit()



#     # page = Page('https://pythonprogramming.net/parsememcparseface/')
#     # soup = bs.BeautifulSoup(page.html, 'html.parser')
#     # js_test = soup.find('p', class_='jstest')

# print('sd')
# URL = 'https://in.tradingview.com/symbols/NSE-MRPL/'
# page = Page(URL)
# soup = bs.BeautifulSoup(page.html, 'lxml')

# # soup = bs.BeautifulSoup(source, 'lxml')

# # soup = BeautifulSoup(r.content, 'html5lib') 
# # print(soup)
# table = soup.find('div', attrs = {'class':'tv-widget-fundamentals__item'}) 

# print(table)
# print(type(soup))

# # table = soup.find_all('table', attrs = {'id':'equityInfo'})
# # print(table)
# '''

# import openpyxl

# path = "C:\\Users\\NishchalMN\\Desktop\\stock\\book1.xlsx"
  
# wb = openpyxl.load_workbook(path) 
  
# sheet = wb.active 

# cell_obj = sheet.cell(row = 11, column = 16) 

# cell_obj.value = 8.45

# wb.save("C:\\Users\\NishchalMN\\Desktop\\stock\\book1.xlsx")

from pprint import pprint
import bs4 as bs
import openpyxl
import sys
import time
from PyQt5.QtWebEngineWidgets import QWebEnginePage
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QUrl
from PyQt5 import QtCore, QtWidgets, QtWebEngineWidgets
print('loaded all modules')

class WebPage(QWebEnginePage):
    def __init__(self):
        super(WebPage, self).__init__()
        self.loadFinished.connect(self.handleLoadFinished)
        self.html = ''

    def company(self):
        soup = bs.BeautifulSoup(self.html, 'lxml')
        #-----------------------------------------------------------------------------
        # EPS, MarketCap, DividendYield, P.E
        table = soup.find_all('div', attrs = {'class':'tv-category-header__fundamentals js-header-fundamentals'}) 
        d1 = {}  # dictionary having EPS, MarketCap, DividendYield, P.E, SharePrice

        try:
            for i in table[0].contents[1:]:
                d1[i.contents[3].contents[0]] = i.contents[1].contents[0]

        except:
            print('Error while calculating EPS, Div, PE !')
        # pprint(d1)
        #------------------------------------------------------------------------------

        # Share Price
        table = soup.find_all('div', attrs = {'class':'tv-symbol-price-quote__value js-symbol-last'})

        try:
            d1['SharePrice'] = table[0].contents[0].contents[0]

        except:
            print('Error while calculating Share price !')

        # pprint(d1)
        #------------------------------------------------------------------------------

        table = soup.find_all('div', attrs = {'class':'tv-widget-fundamentals tv-widget-fundamentals--card-view'})
        d2 = {}
        table = table[0].contents
        pprint(table) 
        try:
            for i in table:
                try:
                    for j in i.contents:
                        try:
                            d2[j.span.contents[0].strip()] = j.span.nextSibling.nextSibling.contents[0].strip()
                        except:
                            pass
                except:
                    pass
        except:
            print('Error while calculating all the values !')

        # pprint(d2)

        #--------------------------------------------------------------

        # trying for custom values

        try:
            MCS = float(d2['Market Capitalization'][:len(d2['Market Capitalization'])-1]) / float(d2['Total Revenue (FY)'][:len(d2['Total Revenue (FY)'])-1])
            d2['MCS'] = MCS
        except:
            d2['MCS'] = '—'
            print('Not enough values for Market value/Sales')

        try:
            ES = float(d2['Enterprise Value (MRQ)'][:len(d2['Enterprise Value (MRQ)'])-1]) / float(d2['Total Revenue (FY)'][:len(d2['Total Revenue (FY)'])-1])
            d2['ES'] = ES
        except:
            d2['ES'] = '—'
            print('Not enough values for EV/Sales')

        pprint(d1)
        pprint(d2)
        # return (d1, d2)

    def start(self, urls):
        self._urls = iter(urls)
        self.fetchNext()

    def fetchNext(self):
        try:
            url = next(self._urls)
            print(url)
            self.load(QUrl(url))
            print('tttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttttt')
           
            self.triggerAction(QWebEnginePage.ReloadAndBypassCache)
            self.triggerAction(QWebEnginePage.Forward, True)
            print('sssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss')
        except StopIteration:
            return False
        # else:
            
        return True

    def processCurrentPage(self, html_str):
        url = self.url().toString()
        print('loaded: [%d chars] %s' % (len(html_str), url))
        self.html = html_str
        self.company()
        time.sleep(3)
        print('completed: [%d chars] %s' % (len(html_str), url))
        
        if not self.fetchNext():
            QtWidgets.qApp.quit()

    def handleLoadFinished(self):
        url = self.url().toString()
        print('url is -------------------' + url)
        self.toHtml(self.processCurrentPage)



urls = ['https://in.tradingview.com/symbols/NSE-MRPL', 'https://in.tradingview.com/symbols/NSE-ONGC']
# print('getting for '+ URL)
app = QtWidgets.QApplication(sys.argv)
webpage = WebPage()
webpage.start(urls)
sys.exit(app.exec_())
print('came')
#----------------------------------------------------------------------------------------------------
# work on excel

# import openpyxl

# path = "C:\\Users\\NishchalMN\\Desktop\\stock\\book1.xlsx"
  
# wb = openpyxl.load_workbook(path) 
  
# sheet = wb.active 

# d1 = {}
# d2 = {}

# req = ['SharePrice', 'P/E', 'EPS', 'Price to Book (FY)', 'Div Yield', 'Enterprise Value/EBITDA (TTM)', 'MCS', 'ES', 'Market Capitalization', '52 Week High', '52 Week Low']


# for comp in range(12, 21):
#     # try:
#     name = sheet.cell(row = comp, column = 2)
#     d1, d2 = company(name.value)
#     d = [d1, d1, d1, d2, d1, d2, d2, d2, d2, d2, d2]
#     print('sddddd')
#     print(d1, d2)
#     for i in range(16, 27):
#         try:
#             cell = sheet.cell(row = comp, column = i) 
#             cell.value = d[i-16][req[i-16]]
#         except:
#             print("Couldn't get " + d[i-16][req[i-16]] + " value")
#     # except:
#     #     print('Company code ' + sheet.cell(row = comp, column = 2).value + ' not found')

# wb.save("C:\\Users\\NishchalMN\\Desktop\\stock\\book1.xlsx")

# print('done')