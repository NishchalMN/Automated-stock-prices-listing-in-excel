from pprint import pprint
import bs4 as bs
import openpyxl
import sys
from PyQt5.QtWebEngineWidgets import QWebEnginePage
from PyQt5.QtWidgets import QApplication
from PyQt5.QtCore import QUrl
print('loaded all modules')


class Page(QWebEnginePage):
    def __init__(self, url):
        self.app = QApplication(sys.argv)
        QWebEnginePage.__init__(self)
        self.html = ''
        self.loadFinished.connect(self.on_load_finished)
        self.load(QUrl(url))
        self.triggerAction(QWebEnginePage.ReloadAndBypassCache, True)
        self.app.exec_()

    def on_load_finished(self):
        self.html = self.toHtml(self.Callable)
        print('Load finished')

    def Callable(self, html_str):
        self.html = html_str
        self.app.quit()

def company(id):
    URL = 'https://in.tradingview.com/symbols/NSE-' + id
    print('getting for '+ URL)
    page = Page(URL)
    print('came')
    soup = bs.BeautifulSoup(page.html, 'html5lib')

    #-----------------------------------------------------------------------------
    # EPS, MarketCap, DividendYield, P.E
    table = soup.find_all('div', attrs = {'class':'tv-category-header__fundamentals js-header-fundamentals'}) 
    d1 = {}  # dictionary having EPS, MarketCap, DividendYield, P.E, SharePrice

    try:
        for i in table[0].contents[1:]:
            d1[i.contents[3].contents[0]] = i.contents[1].contents[0]

    except:
        print('Error while calculating EPS, Div, PE !')
    # pprint(d1)
    #------------------------------------------------------------------------------

    # Share Price
    table = soup.find_all('div', attrs = {'class':'tv-symbol-price-quote__value js-symbol-last'})

    try:
        d1['SharePrice'] = table[0].contents[0].contents[0]

    except:
        print('Error while calculating Share price !')

    # pprint(d1)
    #------------------------------------------------------------------------------

    table = soup.find_all('div', attrs = {'class':'tv-widget-fundamentals tv-widget-fundamentals--card-view'})
    d2 = {}
    table = table[0].contents
    # print(table) 
    try:
        for i in table:
            try:
                for j in i.contents:
                    try:
                        d2[j.span.contents[0].strip()] = j.span.nextSibling.nextSibling.contents[0].strip()
                    except:
                        pass
            except:
                pass
    except:
        print('Error while calculating all the values !')

    # pprint(d2)

    #--------------------------------------------------------------

    # trying for custom values

    try:
        MCS = float(d2['Market Capitalization'][:len(d2['Market Capitalization'])-1]) / float(d2['Total Revenue (FY)'][:len(d2['Total Revenue (FY)'])-1])
        d2['MCS'] = MCS
    except:
        d2['MCS'] = '—'
        print('Not enough values for Market value/Sales')

    try:
        ES = float(d2['Enterprise Value (MRQ)'][:len(d2['Enterprise Value (MRQ)'])-1]) / float(d2['Total Revenue (FY)'][:len(d2['Total Revenue (FY)'])-1])
        d2['ES'] = ES
    except:
        d2['ES'] = '—'
        print('Not enough values for EV/Sales')

    pprint(d1)
    pprint(d2)
    return (d1, d2)
#----------------------------------------------------------------------------------------------------
# work on excel

import openpyxl

path = "C:\\Users\\NishchalMN\\Desktop\\stock\\book1.xlsx"
  
wb = openpyxl.load_workbook(path) 
  
sheet = wb.active 

d1 = {}
d2 = {}

req = ['SharePrice', 'P/E', 'EPS', 'Price to Book (FY)', 'Div Yield', 'Enterprise Value/EBITDA (TTM)', 'MCS', 'ES', 'Market Capitalization', '52 Week High', '52 Week Low']


for comp in range(12, 21):
    # try:
    name = sheet.cell(row = comp, column = 2)
    d1, d2 = company(name.value)
    d = [d1, d1, d1, d2, d1, d2, d2, d2, d2, d2, d2]
    print('sddddd')
    print(d1, d2)
    for i in range(16, 27):
        try:
            cell = sheet.cell(row = comp, column = i) 
            cell.value = d[i-16][req[i-16]]
        except:
            print("Couldn't get " + d[i-16][req[i-16]] + " value")
    # except:
    #     print('Company code ' + sheet.cell(row = comp, column = 2).value + ' not found')

wb.save("C:\\Users\\NishchalMN\\Desktop\\stock\\book1.xlsx")

print('done')