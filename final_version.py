from pprint import pprint

import openpyxl
import sys
from selenium import webdriver

from bs4 import BeautifulSoup
import bs4 as bs

print('loaded all modules')

def company(id, exchange):

    URL = 'https://in.tradingview.com/symbols/' + exchange + '-' + id
    print('getting for '+ URL)
    
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    driver = webdriver.Chrome(chrome_options = options)

    driver.get(URL)

    html = driver.page_source
    
    soup = bs.BeautifulSoup(html, 'html5lib')
    # print(soup)


    #-----------------------------------------------------------------------------
    # EPS, MarketCap, DividendYield, P.E
    table = soup.find_all('div', attrs = {'class':'tv-category-header__fundamentals js-header-fundamentals'}) 
    d1 = {}  # dictionary having EPS, MarketCap, DividendYield, P.E, SharePrice

    try:
        for i in table[0].contents[1:]:
            d1[i.contents[3].contents[0]] = i.contents[1].contents[0]

    except:
        print('Error while calculating EPS, Div, PE !')
    # pprint(d1)



    #------------------------------------------------------------------------------
    # Share Price
    table = soup.find_all('div', attrs = {'class':'tv-symbol-price-quote__value js-symbol-last'})

    try:
        d1['SharePrice'] = table[0].contents[0].contents[0]

    except:
        print('Error while calculating Share price !')

    # pprint(d1)



    #------------------------------------------------------------------------------
    # All other values

    table = soup.find_all('div', attrs = {'class':"tv-feed-widget__scroll-content js-scroll-content"})
    d2 = {}
    table = table[0].div.contents
    # pprint(table) 

    try:
        for i in table:
            try:
                for j in i.contents:
                    try:
                        d2[j.span.contents[0].strip()] = j.span.nextSibling.nextSibling.contents[0].strip()
                    except:
                        pass
            except:
                pass
    except:
        print('Error while calculating all the values !')

    # pprint(d2)

    #--------------------------------------------------------------

    # for custom values

    try:
        MCS = float(d2['Market Capitalization'][:len(d2['Market Capitalization'])-1]) / float(d2['Total Revenue (FY)'][:len(d2['Total Revenue (FY)'])-1])
        d2['MCS'] = MCS
    except:
        d2['MCS'] = '—'
        print('Not enough values for Market value/Sales')

    try:
        ES = float(d2['Enterprise Value (MRQ)'][:len(d2['Enterprise Value (MRQ)'])-1]) / float(d2['Total Revenue (FY)'][:len(d2['Total Revenue (FY)'])-1])
        d2['ES'] = ES
    except:
        d2['ES'] = '—'
        print('Not enough values for EV/Sales')

    pprint(d1)
    pprint(d2)

    return (d1, d2)



#----------------------------------------------------------------------------------------------------
# work on excel

import openpyxl

file = open('values.txt', 'r')
lines = file.readlines()

path = eval(lines[0].split('=')[1])           # path of the excel file
  
wb = openpyxl.load_workbook(path) 
  
sheet = wb.active 

req = eval(lines[1].split('=')[1])             # required stock fields

failed = []

# trying using NSE
for comp in range(eval(lines[2].split('=')[1])[0], eval(lines[3].split('=')[1])[0]):             # name of the stock row number
    try:
        name = sheet.cell(row = comp, column = eval(lines[2].split('=')[1])[1])                   # name of the stock column number
        d1, d2 = company(name.value, 'NSE')

        # print(d1, d2)
        custom = eval(lines[4].split('=')[1])[1]            # custom cell column number
        for i in range(custom, custom + len(req)):         
            try:
                cell = sheet.cell(row = comp, column = i) 
                temp = d2[req[i-custom]] if(req[i-custom] in d2) else d1[req[i-custom]]
                cell.value = temp
            except:
                print("Couldn't get " + req[i-custom] + " value")
    except:
        print('Company code ' + sheet.cell(row = comp, column = eval(lines[2].split('=')[1])[1]).value + ' not found')
        failed.append((sheet.cell(row = comp, column = eval(lines[2].split('=')[1])[1]).value, comp))



print(failed)

# trying using BSE
for (name, comp) in failed:
    try:
        d1, d2 = company(name, 'BSE')

        # print(d1, d2)
        custom = eval(lines[4].split('=')[1])[1]
        for i in range(custom, custom+len(req)):
            try:
                cell = sheet.cell(row = comp, column = i) 
                temp = d2[req[i-custom]] if(req[i-custom] in d2) else d1[req[i-custom]]
                cell.value = temp
            except:
                print("Couldn't get " + req[i-custom] + " value")
    except:
        print('Company code ' + sheet.cell(row = comp, column = eval(lines[2].split('=')[1])[1]).value + ' not found')

wb.save(eval(lines[5].split('=')[1]))

print('done')
# 'SharePrice', '52 Week High', '52 Week Low','Market Capitalization', 'Enterprise Value (MRQ)', 'EPS', 'P/E', 'Price to Book (FY)','Price to Revenue Ratio (TTM)', 'Div Yield', 'Return on Equity (TTM)', 'Return on Assets (TTM)','Return on Invested Capital (TTM)','Revenue per Employee (TTM)', 'EBITDA (TTM)', 'Last Year Revenue (FY)', 'Enterprise Value/EBITDA (TTM)', 'Market Cap/Sales', 'Enterprise Value/Sales'